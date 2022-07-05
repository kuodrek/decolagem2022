import openpyxl as xl

def get_data():
    workbook = xl.load_workbook(filename="dadoscompletosMONOPLANO.xlsx", data_only=True)
    vars_dados = workbook['vars']

    # Variáveis globais
    timestep = vars_dados.cell(row=3,column=4).value
    headwind = vars_dados.cell(row=4,column=4).value
    TOW = vars_dados.cell(row=5,column=4).value
    n_helice = vars_dados.cell(row=6,column=4).value
    altitude = vars_dados.cell(row=7,column=4).value
    mi = vars_dados.cell(row=8,column=4).value
    Sd = vars_dados.cell(row=9,column=4).value
    ac_eh = vars_dados.cell(row=10,column=4).value

    # Variáveis da asa
    Sw = vars_dados.cell(row=3,column=5).value
    iw = vars_dados.cell(row=3,column=6).value
    CLmax_w = vars_dados.cell(row=3,column=7).value
    CL_0_w = vars_dados.cell(row=3,column=8).value
    CL_alfa_w = vars_dados.cell(row=3,column=9).value
    CD_fus = vars_dados.cell(row=3,column=10).value
    Cm_ac_w = vars_dados.cell(row=3,column=11).value
    MAC_w = vars_dados.cell(row=3,column=12).value
    ## Dados para polyfit das curvas de arrasto parasita e induzido
    v1_w = vars_dados.cell(row=5,column=5).value
    v2_w = vars_dados.cell(row=6,column=5).value
    CDp_w1 = vars_dados.cell(row=5,column=6).value
    CDp_w2 = vars_dados.cell(row=6,column=6).value

    a_1_w = vars_dados.cell(row=5,column=7).value
    a_2_w = vars_dados.cell(row=6,column=7).value
    a_3_w = vars_dados.cell(row=7,column=7).value
    a_4_w = vars_dados.cell(row=8,column=7).value
    a_5_w = vars_dados.cell(row=9,column=7).value
    a_6_w = vars_dados.cell(row=10,column=7).value

    CDiw_1 = vars_dados.cell(row=5,column=8).value
    CDiw_2 = vars_dados.cell(row=6,column=8).value
    CDiw_3 = vars_dados.cell(row=7,column=8).value
    CDiw_4 = vars_dados.cell(row=8,column=8).value
    CDiw_5 = vars_dados.cell(row=9,column=8).value
    CDiw_6 = vars_dados.cell(row=10,column=8).value

    # Variáveis do estabilizador horizontal (EH)
    Sh = vars_dados.cell(row=3,column=13).value
    ih = vars_dados.cell(row=3,column=14).value
    CL_0_h = vars_dados.cell(row=3,column=15).value
    CL_alfa_h = vars_dados.cell(row=3,column=16).value
    CL_max_h = vars_dados.cell(row=3,column=17).value
    de_decolagem = vars_dados.cell(row=3,column=18).value
    taue = vars_dados.cell(row=3,column=19).value
    epsilon_0 = vars_dados.cell(row=3,column=20).value
    depsilon_dalfa  = vars_dados.cell(row=3,column=21).value

    ## Dados para polyfit das curvas de arrasto parasita e induzido
    v1_h = vars_dados.cell(row=5,column=13).value
    v2_h = vars_dados.cell(row=6,column=13).value
    CDp_h1 = vars_dados.cell(row=5,column=14).value
    CDp_h2 = vars_dados.cell(row=6,column=14).value

    a_1_h = vars_dados.cell(row=5,column=15).value
    a_2_h = vars_dados.cell(row=6,column=15).value
    a_3_h = vars_dados.cell(row=7,column=15).value
    a_4_h = vars_dados.cell(row=8,column=15).value
    a_5_h = vars_dados.cell(row=9,column=15).value
    a_6_h = vars_dados.cell(row=10,column=15).value

    CDih_1 = vars_dados.cell(row=5,column=16).value
    CDih_2 = vars_dados.cell(row=6,column=16).value
    CDih_3 = vars_dados.cell(row=7,column=16).value
    CDih_4 = vars_dados.cell(row=8,column=16).value
    CDih_5 = vars_dados.cell(row=9,column=16).value
    CDih_6 = vars_dados.cell(row=10,column=16).value

    CL_de = vars_dados.cell(row=6,column=18).value
    Cm_de = vars_dados.cell(row=6,column=19).value
    CL_q = vars_dados.cell(row=6,column=20).value
    Cm_q = vars_dados.cell(row=6,column=21).value
    
def get_data_teste():
    dados_planilha = {
        'de_takeoff': -5,
        'ac_eh': 40,
        'Sd': 50,
        'g': 9.81,
        'rho': 1.225,
        'm': 15,
        'x_tdp': 1,
        'x_tdn': 1,
        'Iyy': 0.5,
        'dt': 0,
        'Sref': 1,
        'mi': 0.05,
        'CL_alfa': 0.09,
        'CL_de': 0.05,
        'CL_q': 0.09,
        'CL_0': 0.9,
        'Cm_alfa': -0.09,
        'Cm_de': -0.05,
        'Cm_q': -0.3,
        'Cm_0': 0.05,
    }
    return dados_planilha

    # CL_alfa = dados_planilha['CL_alfa']
    # CL_de = dados_planilha['CL_de']
    # CL_q = dados_planilha['CL_q']
    # CL_0 = dados_planilha['CL_0']
    # Cm_alfa = dados_planilha['Cm_alfa']
    # Cm_de = dados_planilha['Cm_de']
    # Cm_q = dados_planilha['Cm_q']
    # Cm_0 = dados_planilha['Cm_0']
    # helice_dados = dados_planilha['n_helice']